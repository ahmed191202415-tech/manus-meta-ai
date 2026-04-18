from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from app.config import EXPORT_DIR
from app.schemas.report_requests import SaveExcelReportRequest
from app.utils.files import sanitize_file_name


def save_excel_report_local(body: SaveExcelReportRequest) -> Dict[str, Any]:
    if not body.sheets:
        return {"success": False, "message": "At least one sheet is required."}

    wb = Workbook()
    first = True

    for sheet_spec in body.sheets:
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = (sheet_spec.name or "Sheet")[:31]
        ws.sheet_view.rightToLeft = True

        row_idx = 1
        if body.title and ws.title == (body.sheets[0].name or "Sheet1")[:31]:
            ws.cell(row=row_idx, column=1, value=body.title)
            ws.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="right")
            row_idx += 2

        header_row_index = row_idx
        for col_idx, header in enumerate(sheet_spec.headers, start=1):
            cell = ws.cell(row=header_row_index, column=col_idx, value=header)
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for values in sheet_spec.rows:
            row_idx += 1
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.freeze_panes = f"A{header_row_index + 1}"

        if sheet_spec.headers and sheet_spec.rows:
            end_col = get_column_letter(len(sheet_spec.headers))
            table_ref = f"A{header_row_index}:{end_col}{header_row_index + len(sheet_spec.rows)}"
            table = Table(displayName=f"Table_{abs(hash(ws.title)) % 100000}", ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)

        widths = {}
        for row in ws.iter_rows(min_row=header_row_index, max_row=ws.max_row):
            for cell in row:
                val = "" if cell.value is None else str(cell.value)
                widths[cell.column] = max(widths.get(cell.column, 10), min(len(val) + 3, 40))

        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    file_name = sanitize_file_name(body.file_name, "report", "xlsx")
    output_path = EXPORT_DIR / file_name
    wb.save(output_path)

    return {
        "success": True,
        "file_name": file_name,
        "file_path": str(output_path),
        "export_dir": str(EXPORT_DIR),
    }
